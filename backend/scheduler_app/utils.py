from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import datetime

def export_timetable_pdf(timetable):
    """Export timetable to PDF format"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=72, leftMargin=72,
                           topMargin=72, bottomMargin=18)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Title'],
        fontSize=16,
        alignment=1,
        spaceAfter=12
    )
    
    story = []
    
    # Title
    title = Paragraph(f"{timetable.name}", title_style)
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Get classes
    classes = timetable.classes.all().select_related(
        'course', 'instructor', 'room', 'meeting_time', 'section'
    )
    
    # Create timetable grid
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    time_slots = []
    
    # Get all unique time slots
    for cls in classes:
        time_slot = f"{cls.meeting_time.start_time.strftime('%H:%M')}-{cls.meeting_time.end_time.strftime('%H:%M')}"
        if time_slot not in time_slots:
            time_slots.append(time_slot)
    
    time_slots.sort()
    
    # Create table data
    table_data = []
    header_row = ['Time'] + days
    table_data.append(header_row)
    
    for time_slot in time_slots:
        row = [time_slot]
        for day in days:
            cell_content = ""
            day_classes = classes.filter(
                meeting_time__day=day,
                meeting_time__start_time__time=datetime.datetime.strptime(time_slot.split('-')[0], '%H:%M').time()
            )
            
            for cls in day_classes:
                if cell_content:
                    cell_content += "\n"
                cell_content += f"{cls.course.course_id}\n{cls.instructor.name}\n{cls.room.room_number}\n{cls.section.section_id}"
            
            row.append(cell_content if cell_content else "")
        
        table_data.append(row)
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    return buffer.getvalue()

def export_timetable_excel(timetable):
    """Export timetable to Excel format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Timetable"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Get classes
    classes = timetable.classes.all().select_related(
        'course', 'instructor', 'room', 'meeting_time', 'section'
    )
    
    # Create timetable grid
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    time_slots = []
    
    # Get all unique time slots
    for cls in classes:
        time_slot = f"{cls.meeting_time.start_time.strftime('%H:%M')}-{cls.meeting_time.end_time.strftime('%H:%M')}"
        if time_slot not in time_slots:
            time_slots.append(time_slot)
    
    time_slots.sort()
    
    # Write header
    ws['A1'] = timetable.name
    ws.merge_cells('A1:G1')
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = center_alignment
    
    # Write column headers
    headers = ['Time'] + days
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Write timetable data
    for row_idx, time_slot in enumerate(time_slots, 4):
        ws.cell(row=row_idx, column=1, value=time_slot).alignment = center_alignment
        
        for col_idx, day in enumerate(days, 2):
            cell_content = ""
            day_classes = classes.filter(
                meeting_time__day=day,
                meeting_time__start_time__time=datetime.datetime.strptime(time_slot.split('-')[0], '%H:%M').time()
            )
            
            for cls in day_classes:
                if cell_content:
                    cell_content += "\n"
                cell_content += f"{cls.course.course_id}\n{cls.instructor.name}\n{cls.room.room_number}\n{cls.section.section_id}"
            
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_content)
            cell.alignment = center_alignment
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = min(adjusted_width, 50)
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()